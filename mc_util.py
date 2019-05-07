# coding: utf-8
import numpy as np
import pandas as pd
import openmatrix as omx
from IPython.display import display
from openpyxl import load_workbook,Workbook
from time import strftime
import os.path
from config import out_path

AO_dict = {'DA':1,'SR2':2,'SR3+':3.5,'SM_RA':1, 'SM_SH':2}

def store_omx_as_dict(infile_path):
	'''
	Given an omx file, it stores the matrices in a dictionary.
	:param infile_path: path of the omx file.
	:returns: dict
	'''
	store_dict = {}
	with omx.open_file(infile_path,'r') as f:
		for name in f.list_matrices():
			store_dict[name] = np.array(f[name])
	return store_dict

	
def expand_prod(var_vector):
	'''
	This returns a 2730 x 2730 table for a zonal variable vector that applies to the production side.
	:param var_vector: either a pandas series or a numpy array with at least 2730 entries
	:returns: numpy array 2730 x 2730
	:raises: ValueError
	'''
	try:
		return np.repeat(var_vector.values[:2730].reshape(2730,1),2730,axis= 1)
	except:
		try: return np.repeat(var_vector[:2730].reshape(2730,1),2730,axis= 1)
		except: raise ValueError('error expanding vector, check size and input type')

def expand_attr(var_vector):
	'''
	This returns a 2730 x 2730 table for a zonal variable vector that applies to the attraction side.
	:param var_vector: either a pandas series or a numpy array with at least 2730 entries
	:returns: numpy array 2730 x 2730
	:raises: ValueError
	'''	

	try:
		return np.repeat(var_vector.values[:2730].reshape(1,2730),2730,axis= 0)
	except:
		try: return np.repeat(var_vector[:2730].reshape(1,2730),2730,axis = 0)
		except: raise ValueError('error expanding vector')
	
	
def write_trip_tables(mc_obj,out_fn):
	'''
	This writes the resulting trip tables of mode choice to a .omx file.
	:param mc_obj: mode choice module object as defined in the IPython notebook
	:param out_fn: path of output .omx file
	'''

	modes = mc_obj.param['mode']
	with omx.open_file(out_fn,'w') as ttmc:
		for pv in mc_obj.peak_veh:
			for mode in modes:
				ttmc[f'{mode}_{pv}'] = mc_obj.trips_by_mode[pv][mode]
	
def display_mode_share(mc_obj):
	'''
	This displays a mode share summary by market segment (with / without vehicle, peak / off-peak) on the IPython notebook.
	:param mc_obj: mode choice module object as defined in the IPython notebook
	'''
	modes = mc_obj.param['mode']
	# display mode share tables
	avg_trips_by_mode = pd.DataFrame(index=mc_obj.peak_veh,columns = modes)
	for pv in mc_obj.peak_veh:
		for mode in modes:
			avg_trips_by_mode.loc[pv,mode] = mc_obj.trips_by_mode[pv][mode].sum()
	avg_mode_share = avg_trips_by_mode.divide(avg_trips_by_mode.sum(1),axis = 0)

	display(avg_mode_share.style.format("{:.2%}"))
	
	
def write_mode_share_to_excel(mc_obj,purpose, out_excel_fn = out_path + "MC_mode_share_{}.xlsx".format(strftime("%Y%m%d_%H%M%S"))):
	'''
	Writes mode share summary by purpose and market segment to an Excel workbook.
	:param mc_obj: mode choice module object as defined in the IPython notebook
	:param purpose: can be a single purpose or 'all', in which case the Excel workbook has six sheets, one for each purpose.
	:param out_excel_fn: output Excel filename, by default in the output path defined in config.py	
	'''

	if purpose == 'all':
		# check if file exists.
		if os.path.isfile(out_excel_fn):
			book = load_workbook(out_excel_fn)
		else: 
			book = Workbook()
			book.save(out_excel_fn)
		writer = pd.ExcelWriter(out_excel_fn,engine = 'openpyxl')
		writer.book = book
		
		for purp in ['HBW','HBO', 'NHB', 'HBSc1', 'HBSc2', 'HBSc3']:
			trip_table = mc_obj.table_container.get_table(purp)
			mode_share = pd.DataFrame(columns = mc_obj.peak_veh)
			for pv in mc_obj.peak_veh:
				for mode in trip_table[pv].keys():
					mode_share.loc[mode,pv] = trip_table[pv][mode].sum()
			
			mode_share['Total'] = mode_share.sum(1)
			mode_share['Share'] = mode_share['Total'] / mode_share['Total'].sum()
			
			if purp in book.sheetnames: # if sheetname exists, delete
				book.remove(book[purp])
				writer.save()

			mode_share.to_excel(writer, sheet_name = purp)
		
			writer.save()
		
	elif purpose in ['HBW','HBO', 'NHB', 'HBSc1', 'HBSc2', 'HBSc3']:
		# check if file exists.
		if os.path.isfile(out_excel_fn):
			book = load_workbook(out_excel_fn)
		else: 
			book = Workbook()
			book.save(out_excel_fn)
		writer = pd.ExcelWriter(out_excel_fn,engine = 'openpyxl')
		writer.book = book
		
		mode_share = pd.DataFrame(columns = mc_obj.peak_veh)
		for pv in mc_obj.peak_veh:
			for mode in mc_obj.trips_by_mode[pv].keys():
				mode_share.loc[mode,pv] = mc_obj.trips_by_mode[pv][mode].sum()
		
		mode_share['Total'] = mode_share.sum(1)
		mode_share['Share'] = mode_share['Total'] / mode_share['Total'].sum()
		
		if purpose in book.sheetnames: # if sheetname exists, delete
			book.remove(book[purpose])
			writer.save()

		mode_share.to_excel(writer, sheet_name = purpose)
	
		writer.save()
		
def mt_prod_attr_nhood(mc_obj, trip_table, skim): # miles traveled. For VMT and PMT, by neighborhood
	# sum prodct of trip_table - skims
	mt_total = trip_table * skim['Length (Skim)'] 
	
	# calculate marginals
	prod = pd.DataFrame(np.sum(mt_total,axis = 1)/2, columns = ['Production'])
	attr = pd.DataFrame(np.sum(mt_total,axis = 0) / 2, columns = ['Attraction'])
	
	towns = mc_obj.taz_lu.sort_values('TAZ_ID').iloc[0:2730]
	mt_taz = pd.concat([towns[['TAZ_ID','BOSTON_NB']],prod,attr],axis = 1,join = 'inner')
	mt_taz.index.names=['Boston Neighborhood']
	return mt_taz.groupby(['BOSTON_NB']).sum()[['Production','Attraction']].reset_index()
	
def vmt_by_neighborhood(mc_obj, out_fn = None, by = None):
	'''
	Summarizes VMT production and attraction by the 26 Boston neighborhoods.
	:param mc_obj: mode choice module object as defined in the IPython notebook
	:param out_fn: output csv filename; if None specified, in the output path defined in config.py	
	:param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
	'''
	
	if out_fn is None and by is None:
		out_fn = out_path + f'vmt_by_neighborhood.csv'
	elif out_fn is None and by:
		out_fn = out_path + f'vmt_by_neighborhood_by_{by}.csv'
	skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
	if by in ['peak','veh_own','purpose'] == False:
		print('Only supports VMT by neighborhood, peak / vehicle ownership, purpose.')
		return
		
	else:
		vmt_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
		for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
			for peak in ['PK','OP']:
				for veh_own in ['0','1']:
					if mc_obj.table_container.get_table(purpose):
						drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
						
						auto_trip_table = sum([
						mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] / AO_dict[mode]
 for mode in ['DA','SR2','SR3+'] if mode in drive_modes])
						
						
						
						vmt_table = mt_prod_attr_nhood(mc_obj,auto_trip_table,skim_dict[peak])
						vmt_table['peak'] = peak
						vmt_table['veh_own'] = veh_own
						vmt_table['purpose'] = purpose
						vmt_master_table = vmt_master_table.append(vmt_table)
		
		if by == None:
			vmt_summary = vmt_master_table.groupby('BOSTON_NB').sum()
		
		elif by == 'peak':
			vmt_summary = pd.concat([
			vmt_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
		
		elif by == 'veh_own':
			vmt_summary = pd.concat([
			vmt_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
			)
			
		elif by == 'purpose':
			vmt_summary = pd.concat([
			vmt_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in vmt_master_table.purpose.unique()],axis = 1, keys= vmt_master_table.purpose.unique())

		vmt_summary.to_csv(out_fn)

def pmt_by_neighborhood(mc_obj, out_fn = None, by = None):
	'''
	Summarizes PMT production and attraction by the 26 Boston neighborhoods.
	:param mc_obj: mode choice module object as defined in the IPython notebook
	:param out_fn: output csv filename; if None specified, in the output path defined in config.py	
	:param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
	'''
	
	if out_fn is None and by is None:
		out_fn = out_path + f'pmt_by_neighborhood.csv'
	elif out_fn is None and by:
		out_fn = out_path + f'pmt_by_neighborhood_by_{by}.csv'
		
	skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
	
	if by in ['peak','veh_own','purpose'] == False:
		print('Only supports PMT by neighborhood, peak / vehicle ownership, purpose.')
		return
		
	else:
		pmt_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
		for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
			for peak in ['PK','OP']:
				for veh_own in ['0','1']:
					if mc_obj.table_container.get_table(purpose):
						drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
						person_trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] for mode in ['DA','SR2','SR3+','SM_RA','SM_SH'] if mode in drive_modes])
						
						pmt_table = mt_prod_attr_nhood(mc_obj,person_trip_table,skim_dict[peak])
						pmt_table['peak'] = peak
						pmt_table['veh_own'] = veh_own
						pmt_table['purpose'] = purpose
						pmt_master_table = pmt_master_table.append(pmt_table)
		
		if by == None:
			pmt_summary = pmt_master_table.groupby('BOSTON_NB').sum()
		
		elif by == 'peak':
			pmt_summary = pd.concat([
			pmt_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
		
		elif by == 'veh_own':
			pmt_summary = pd.concat([
			pmt_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
			)
			
		elif by == 'purpose':
			pmt_summary = pd.concat([
			pmt_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in pmt_master_table.purpose.unique()],axis = 1, keys= pmt_master_table.purpose.unique())

		pmt_summary.to_csv(out_fn)

def mode_share_by_neighborhood(mc_obj, out_fn = None, by = None):
	'''
	Summarizes mode share as the average of trips to/from the 26 Boston neighborhoods, in three categories - drive, non-motorized and transit.
	:param mc_obj: mode choice module object as defined in the IPython notebook
	:param out_fn: output csv filename; if None specified, in the output path defined in config.py	
	:param by: grouping used for the summary
	'''
	if out_fn is None and by is None:
		out_fn = out_path + f'mode_share_by_neighborhood.csv'
	elif out_fn is None and by:
		out_fn = out_path + f'mode_share_by_neighborhood_by_{by}.csv'

	if by in ['peak','veh_own','purpose'] == False:
		print('Only supports PMT by neighborhood, peak / vehicle ownership, purpose.')
		return

	else:
		mode_categories = {'DA':'drive','SR2':'drive','SR3+':'drive','Bike':'non-motorized','Walk':'non-motorized',
	'WAT':'transit','DAT_CR':'transit','DAT_B':'transit','DAT_LB':'transit','DAT_RT':'transit','SM_RA':'smart mobility','SM_SH':'smart mobility'}
		share_master_table = pd.DataFrame(columns = ['drive','non-motorized','transit','peak','veh_own','purpose'])

		for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
			for peak in ['PK','OP']:
				for veh_own in ['0','1']:
					if mc_obj.table_container.get_table(purpose):
						share_table = pd.DataFrame(index = range(0,2730),columns = ['drive','non-motorized','transit']).fillna(0)
						for mode in mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}']:
						
							trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode]
							category = mode_categories[mode]
							
							share_table[category] += (trip_table.sum(axis = 1)+trip_table.sum(axis = 0))/2
						
						towns = mc_obj.taz_lu.sort_values('TAZ_ID').iloc[0:2730]
						trips = pd.concat([towns[['TAZ_ID','BOSTON_NB']],share_table],axis = 1,join = 'inner').groupby(['BOSTON_NB']).sum().drop(['TAZ_ID'],axis = 1)
						trips['peak'] = peak
						trips['veh_own'] = veh_own
						trips['purpose'] = purpose
						
						share_master_table = share_master_table.append(trips.reset_index())
		
		if by == None:
			trip_summary = share_master_table.groupby('BOSTON_NB').sum()
			share_summary = trip_summary.divide(trip_summary.sum(axis = 1),axis = 0)
		
		elif by == 'peak':
			share_summary = pd.concat([
			share_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak].divide(
			share_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak].sum(axis=1),axis = 0)
			for peak in ['PK','OP']
			], axis = 1, keys = ['PK','OP'])
		
		elif by == 'veh_own':
			share_summary = pd.concat([
			share_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own].divide(
			share_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own].sum(axis=1),axis = 0)
			for veh_own in ['0','1']
			], axis = 1, keys = ['No car', 'With car'])
			
		elif by == 'purpose':
			share_summary = pd.concat([
			share_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose].divide(
			share_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose].sum(axis=1),axis = 0)
			for purpose in share_master_table.purpose.unique()
			],axis = 1, keys= share_master_table.purpose.unique())
		
		
		share_summary.to_csv(out_fn)
	
	
def sum_unequal_length(a,b):
    if len(a)>len(b):
        b.resize(a.shape, refcheck=False)
    else:
        a.resize(b.shape, refcheck=False)
    return a+b	
	
def compute_summary_by_subregion(mc_obj,taz_fn = r"X:\Proj\2018\180021 - ISE_GHGMitAnly\001\Model\Mode Choice\TAZ_by_interstate.csv", metric = 'VMT',subregion = 'neighboring'):
	''' Computing function used by write_summary_by_subregion(), does not produce outputs'''

	if metric.lower() not in ('vmt','pmt','mode share'):
		print('Only supports VMT, PMT and mode share calculations.')
		return
	if subregion.lower() not in ('neighboring','i93','i495','region'):
		print('Only supports "neighboring" for towns neighboring Boston, I93, I495 or Region.')
		return
	
	taz = pd.read_csv(taz_fn).sort_values(['ID_FOR_CS']).reset_index().drop(columns = ['index'])[0:2730][['TOWN','in_i95i93','in_i495']]
	taz['BOS_AND_NEI'] = taz['TOWN'].isin([n+',MA' for n in ['WINTHROP','CHELSEA','REVERE','SOMERVILLE','CAMBRIDGE','WATERTOWN','NEWTON',
              'BROOKLINE','NEEDHAM','DEDHAM','MILTON','QUINCY','BOSTON']])
	subregion_dict = {'neighboring':'BOS_AND_NEI','i93':'in_i95i93','i495':'in_i495'}
	
			
	if metric.lower() == 'vmt':
		skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
		vmt_table = np.zeros((2730,2730))
		for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
			for peak in ['PK','OP']:
				for veh_own in ['0','1']:
					if mc_obj.table_container.get_table(purpose):
						drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
						trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] / AO_dict[mode] for mode in ['DA','SR2','SR3+','SM_RA','SM_SH'] if mode in drive_modes])
						vmt_table += trip_table * skim_dict[peak]['Length (Skim)']
						
		if subregion.lower() in subregion_dict:
			field = subregion_dict[subregion.lower()]
			boston_o_auto_vmt = vmt_table[taz['TOWN']=='BOSTON,MA',:][:, taz[field]== True]
			boston_d_auto_vmt = vmt_table[taz[field]== True,:][:,taz['TOWN']=='BOSTON,MA']
			town_definition = taz[taz[field]== True]
		
		elif subregion.lower() == 'region':
			boston_o_auto_vmt = vmt_table[taz['TOWN']=='BOSTON,MA',:]
			boston_d_auto_vmt = vmt_table[:][:,taz['TOWN']=='BOSTON,MA']
			town_definition = taz	
		
		zone_vmt_daily_o = pd.DataFrame(sum_unequal_length(np.sum(boston_o_auto_vmt,axis=0), np.sum(boston_o_auto_vmt,axis=1))/2 ,columns=["VMT"])
		zone_vmt_daily_d = pd.DataFrame(sum_unequal_length(np.sum(boston_d_auto_vmt,axis=0), np.sum(boston_d_auto_vmt,axis=1))/2 ,columns=["VMT"])

		boston_within_auto_vmt = vmt_table[taz['TOWN']=='BOSTON,MA',:][:,taz['TOWN']=='BOSTON,MA']
		zone_vmt_daily_within = pd.DataFrame(np.sum(boston_within_auto_vmt,axis=0)/2+np.sum(boston_within_auto_vmt,axis=1)/2,columns = ['VMT'])

		town_vmt_o=pd.concat([town_definition,zone_vmt_daily_o],axis=1,join='inner')
		town_vmt_d=pd.concat([town_definition,zone_vmt_daily_d],axis=1,join='inner')
		town_vmt_within = pd.concat([taz[taz['TOWN']=='BOSTON,MA'],zone_vmt_daily_within],axis=1,join='inner')

		vmt_sum_o = town_vmt_o[town_vmt_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['VMT']
		vmt_sum_d = town_vmt_d[town_vmt_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['VMT']
		vmt_sum_within = town_vmt_within.groupby(['TOWN']).sum()['VMT']

		boston_portion_vmt = (vmt_sum_o + vmt_sum_d - vmt_sum_within).values[0]
		
		return boston_portion_vmt
	
	elif metric.lower() == 'pmt':
		skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
		
		
		pmt_table = np.zeros((2730,2730))
		for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
			for peak in ['PK','OP']:
				for veh_own in ['0','1']:
					if mc_obj.table_container.get_table(purpose):
						drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
						
						trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] for mode in ['DA','SR2','SR3+'] if mode in drive_modes])
						
						pmt_table += trip_table * skim_dict[peak]['Length (Skim)']
		
		if subregion.lower() in subregion_dict:
			field = subregion_dict[subregion.lower()]
			boston_o_auto_pmt = pmt_table[taz['TOWN']=='BOSTON,MA',:][:, taz[field]== True]
			boston_d_auto_pmt = pmt_table[taz[field]== True,:][:,taz['TOWN']=='BOSTON,MA']
			town_definition = taz[taz[field]== True]
		
		elif subregion.lower() == 'region':
			boston_o_auto_pmt = pmt_table[taz['TOWN']=='BOSTON,MA',:]
			boston_d_auto_pmt = pmt_table[:][:,taz['TOWN']=='BOSTON,MA']
			town_definition = taz
		
		zone_pmt_daily_o = pd.DataFrame(sum_unequal_length(np.sum(boston_o_auto_pmt,axis=0), np.sum(boston_o_auto_pmt,axis=1))/2 ,columns=["PMT"])
		zone_pmt_daily_d = pd.DataFrame(sum_unequal_length(np.sum(boston_d_auto_pmt,axis=0), np.sum(boston_d_auto_pmt,axis=1))/2 ,columns=["PMT"])

		boston_within_auto_pmt = pmt_table[taz['TOWN']=='BOSTON,MA',:][:,taz['TOWN']=='BOSTON,MA']
		zone_pmt_daily_within = pd.DataFrame(np.sum(boston_within_auto_pmt,axis=0)/2+np.sum(boston_within_auto_pmt,axis=1)/2,columns = ['PMT'])

		town_pmt_o=pd.concat([town_definition,zone_pmt_daily_o],axis=1,join='inner')
		town_pmt_d=pd.concat([town_definition,zone_pmt_daily_d],axis=1,join='inner')
		town_pmt_within = pd.concat([taz[taz['TOWN']=='BOSTON,MA'],zone_pmt_daily_within],axis=1,join='inner')

		pmt_sum_o = town_pmt_o[town_pmt_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['PMT']
		pmt_sum_d = town_pmt_d[town_pmt_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['PMT']
		pmt_sum_within = town_pmt_within.groupby(['TOWN']).sum()['PMT']

		boston_portion_pmt = (pmt_sum_o + pmt_sum_d - pmt_sum_within).values[0]
		
		return boston_portion_pmt

	elif metric.lower() == 'mode share':		
		mode_categories = {'DA':'drive','SR2':'drive','SR3+':'drive','Bike':'non-motorized','Walk':'non-motorized',
	'WAT':'transit','DAT_CR':'transit','DAT_B':'transit','DAT_LB':'transit','DAT_RT':'transit','SM_RA':'smart mobility','SM_SH':'smart mobility'}
		
		share_table = dict(zip(['drive','non-motorized','transit','smart mobility'],[0,0,0,0]))
		subregion_dict = {'neighboring':'BOS_AND_NEI','i93':'in_i95i93','i495':'in_i495'}
		if subregion.lower() in subregion_dict:
			field = subregion_dict[subregion.lower()]

			for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
				for peak in ['PK','OP']:
					for veh_own in ['0','1']:
						if mc_obj.table_container.get_table(purpose):
							for mode in mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}']:
								trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode]
								boston_ii_trips = trip_table[taz['TOWN']=='BOSTON,MA',:][:,taz['TOWN']=='BOSTON,MA'].sum()
								trips = trip_table[taz['TOWN']=='BOSTON,MA',:][:, taz[field]== True].sum() + trip_table[taz[field]== True,:][:,taz['TOWN']=='BOSTON,MA'].sum() - boston_ii_trips
								category = mode_categories[mode]
								share_table[category]+=trips
								
		elif subregion.lower() == 'region':
			for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
				for peak in ['PK','OP']:
					for veh_own in ['0','1']:
						if mc_obj.table_container.get_table(purpose):
							for mode in mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}']:
								trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode]
								boston_ii_trips = trip_table[taz['TOWN']=='BOSTON,MA',:][:,taz['TOWN']=='BOSTON,MA'].sum()
								trips = trip_table[taz['TOWN']=='BOSTON,MA',:][:].sum() + trip_table[:][:,taz['TOWN']=='BOSTON,MA'].sum() - boston_ii_trips
								category = mode_categories[mode]
								share_table[category]+=trips
		# normalize
		return (pd.DataFrame.from_dict(share_table,orient = 'index') / (pd.DataFrame.from_dict(share_table,orient = 'index').sum())).to_dict()[0]
		
		
def write_summary_by_subregion(mc_obj, taz_fn =r"X:\Proj\2018\180021 - ISE_GHGMitAnly\001\Model\Mode Choice\TAZ_by_interstate.csv", out_path = out_path):

	'''
	Summarizes VMT, PMT or mode share by subregions of Massachusetts surrounding Boston (neighboring towns of Boston / within I-93/95 / within I-495).
	:param mc_obj: mode choice module object as defined in the IPython notebook
	:param taz_fn: TAZ file that contains subregion definition
	:param out_path: output path.
	'''

	subregion_dict = dict(zip(['neighboring','i93','i495','region'],['Boston and Neighboring Towns', 'Within I-93/95', 'Within I-495', 'Entire Region']))
	vmt_summary_df = pd.DataFrame(index = subregion_dict.values(), columns = ['VMT to/from Boston'])
	pmt_summary_df = pd.DataFrame(index = subregion_dict.values(), columns = ['PMT to/from Boston'])
	mode_share_df = pd.DataFrame(index = subregion_dict.values(),columns = ['drive','non-motorized','transit','smart mobility'])
	for subregion in subregion_dict:
		vmt_summary_df.loc[subregion_dict[subregion]] = compute_summary_by_subregion(mc_obj,taz_fn, metric = 'VMT',subregion = subregion)
		pmt_summary_df.loc[subregion_dict[subregion]] = compute_summary_by_subregion(mc_obj,taz_fn, metric = 'PMT',subregion = subregion)
		mode_share_df.loc[subregion_dict[subregion]] = compute_summary_by_subregion(mc_obj,taz_fn, metric = 'mode share',subregion = subregion)
	vmt_summary_df.to_csv(out_path + 'vmt_summary_subregions.csv')
	pmt_summary_df.to_csv(out_path + 'pmt_summary_subregions.csv')
	mode_share_df.to_csv(out_path + 'mode_share_summary_subregions.csv')
		
def transit_ridership(mc_obj,MBTA_fn =r"X:\Proj\2018\180021 - ISE_GHGMitAnly\001\Model\Mode Choice\MBTA_coverage.csv",out_path = out_path):
	'''
	Summarizes transit ridership by peak period in cities and towns with MBTA subway service.
	:param mc_obj: mode choice module object as defined in the IPython notebook
	:param mbta_fn: TAZ file that contains MBTA coverage definition
	:param out_path: output path.
	'''

	MBTA_cvg = pd.read_csv(MBTA_fn)
	taz_cvg = mc_obj.taz_lu.merge(MBTA_cvg, how = 'left', on = 'TOWN')
	taz_cvg = taz_cvg[['ID_FOR_CS','subway','TOWN']]
	taz_cvg['covered'] = taz_cvg['subway']==1 # 870 TAZs included.
	ridership = {'PK':0, 'OP':0}
	for purpose in ['HBW','HBO','NHB', 'HBSc1','HBSc2','HBSc3']:
		for peak in ['PK','OP']:
			for veh_own in ['0','1']:
				if mc_obj.table_container.get_table(purpose):
					for mode in set(mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'])&set(['WAT','DAT_B','DAT_CR','DAT_LB','DAT_RT']):
						trips = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode][taz_cvg['covered'],:][:,(taz_cvg['TOWN']=='BOSTON,MA')].sum()
						ridership[peak]+=trips
						
	# calculate ridership
	pd.DataFrame.from_dict({'Ridership':ridership}).to_csv(out_path + 'transit_ridership_summary.csv')
	